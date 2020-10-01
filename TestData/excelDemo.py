'''
Created on 29 Sep 2020

@author: 611104661
'''
import openpyxl

book=openpyxl.load_workbook("C:\\PythonT\\PythonSelFramework\\TestData\\PythonDemo.xlsx")
sheet=book.active
#creating empty dictionary
Dict={}

cell=sheet.cell(row=1,column=1)
#print(cell.value)
#===============================================================================
# #Give value
# sheet.cell(row=2,column=2).value="Shivangi"
# 
# print(sheet.cell(row=2,column=2).value)
# #To get row count
# print(sheet.max_row)
# print(sheet.max_column)
# print(sheet["A3"].value)
#===============================================================================

for i in range(1,sheet.max_row+1):#to get rows
    if sheet.cell(row=i,column=1).value=="TestCase1":
        for j in range(2,sheet.max_column+1):# to get columns
            #print(sheet.cell(row=i,column=j).value)
            # to Store data in Dict
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
            
print(Dict)
    