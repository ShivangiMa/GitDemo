'''
Created on 28 Sep 2020

@author: 611104661
'''
import openpyxl

class HomePageData:

    test_HomePageDemoData= [{"Name":"Shivangi","Email":"Maurya@gmail.com","Gender":"Female"},{"Name":"Krishna","Email":"Kumar@gmail.com","Gender":"Male"}]

    @staticmethod
    def getTestData(testcasename):

        #creating empty dictionary
        Dict={}
        book=openpyxl.load_workbook("C:\\PythonT\\PythonSelFramework\\TestData\\PythonDemo.xlsx")
        sheet=book.active

        for i in range(1,sheet.max_row+1):#to get rows
            if sheet.cell(row=i,column=1).value==testcasename:
                for j in range(2,sheet.max_column+1):# to get columns
                    #print(sheet.cell(row=i,column=j).value)
                    # to Store data in Dict
                    Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value   
        return[Dict]
        #print(Dict)